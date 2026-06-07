"""Phase 4B — report exports (Excel + PDF).

In-process (mongomock). Verifies exports reuse the 4A builders (same totals),
inherit scope guards, and return proper file responses.

Run:  pytest backend/tests/test_phase4b_exports.py
"""
import os
import asyncio
import io
import uuid

import pytest

os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "phase4b_test")
os.environ.setdefault("JWT_SECRET", "test_secret_key_at_least_32_chars_long_xx")

import httpx
from mongomock_motor import AsyncMongoMockClient

import server


def run(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


def auth(t):
    return {"Authorization": f"Bearer {t}"}


class Ctx:
    pass


@pytest.fixture(scope="module")
def ctx():
    server.client = AsyncMongoMockClient()
    server.db = server.client[os.environ["DB_NAME"]]
    transport = httpx.ASGITransport(app=server.app)
    client = httpx.AsyncClient(transport=transport, base_url="http://test")

    async def build():
        c = Ctx()
        c.client = client
        await client.post("/api/seed")
        c.hq = (await client.post("/api/auth/login", json={"email": "admin@dietcafe.com", "password": "admin123"})).json()["token"]

        async def mk_store(n, code):
            return (await client.post("/api/stores", json={"name": n, "code": code}, headers=auth(c.hq))).json()["store_id"]
        c.store_a = await mk_store("A", "AAA")
        c.store_b = await mk_store("B", "BBB")
        c.store_z = await mk_store("Z", "ZZZ")

        async def mk_staff(role, store_id=None, cluster=None, pin="0"):
            body = {"name": role + pin, "role": role, "pin": pin}
            if store_id:
                body["store_id"] = store_id
            if cluster:
                body["cluster_store_ids"] = cluster
            r = await client.post("/api/staff", json=body, headers=auth(c.hq))
            return server.create_token(r.json()["id"], role)
        c.mgr_a = await mk_staff("store_manager", store_id=c.store_a, pin="1111")
        c.cashier_a = await mk_staff("cashier", store_id=c.store_a, pin="2222")
        c.kitchen_a = await mk_staff("kitchen", store_id=c.store_a, pin="3333")
        c.area_ab = await mk_staff("area_manager", cluster=[c.store_a, c.store_b], pin="4444")

        await client.post("/api/admin/migrate-inventory", headers=auth(c.hq))
        c.cust = (await client.post("/api/auth/register", json={"email": "c@t.com", "password": "p", "name": "c", "role": "customer"})).json()["token"]

        # clean raw + product, inward @0.5, then a sale -> non-trivial P&L numbers
        c.prod_id = str(uuid.uuid4())
        await server.db.products.insert_one({
            "id": c.prod_id, "name": "Test Chicken", "product_type": "single",
            "cost_per_100g": 150, "available_qty_grams": 0, "is_active": True,
            "calories_per_100g": 100, "protein_per_100g": 10, "carbs_per_100g": 5, "fat_per_100g": 2})
        await client.post(f"/api/inventory/{c.store_a}/items", headers=auth(c.mgr_a),
                          json={"name": "Test Chicken", "unit": "g", "product_id": c.prod_id, "qty_on_hand": 0, "avg_cost": 0})
        await client.post(f"/api/inventory/{c.store_a}/inward", headers=auth(c.mgr_a),
                          json={"product_id": c.prod_id, "qty": 10000, "purchase_price": 0.5})
        await client.post("/api/orders", headers=auth(c.cust), json={
            "order_type": "dine-in", "store_id": c.store_a,
            "items": [{"product_id": c.prod_id, "product_name": "Test Chicken", "grams": 200, "price": 300,
                       "quantity": 1, "calories": 1, "protein": 1, "carbs": 1, "fat": 1, "product_type": "single"}],
            "total_price": 300, "item_subtotal": 300, "total_calories": 1, "total_protein": 1,
            "total_carbs": 1, "total_fat": 1, "confirm_duplicate": True})
        return c

    c = run(build())
    yield c
    run(client.aclose())


# ---------- file responses ----------

def test_xlsx_and_pdf_content_types(ctx):
    async def go():
        x = await ctx.client.get("/api/reports/pnl/export.xlsx", headers=auth(ctx.mgr_a))
        assert x.status_code == 200, x.text
        assert x.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in x.headers.get("content-disposition", "")
        assert x.content[:2] == b"PK"  # xlsx is a zip
        p = await ctx.client.get("/api/reports/pnl/export.pdf", headers=auth(ctx.mgr_a))
        assert p.status_code == 200, p.text
        assert p.headers["content-type"] == "application/pdf"
        assert p.content[:4] == b"%PDF"
    run(go())


# ---------- exported totals equal the JSON builder ----------

def test_xlsx_totals_match_json(ctx):
    async def go():
        j = (await ctx.client.get("/api/reports/pnl", headers=auth(ctx.mgr_a))).json()
        row_a = next(r for r in j if r["store_id"] == ctx.store_a)
        x = await ctx.client.get("/api/reports/pnl/export.xlsx", headers=auth(ctx.mgr_a))
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(x.content))
        ws = wb["P&L"]
        headers = [c.value for c in ws[1]]
        # find the store_a row
        target = None
        for r in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, r))
            if rec.get("store_id") == ctx.store_a:
                target = rec
                break
        assert target is not None
        assert float(target["revenue"]) == row_a["revenue"] == 300
        assert float(target["cogs"]) == row_a["cogs"] == 100
        assert float(target["gross_profit"]) == row_a["gross_profit"] == 200
    run(go())


# ---------- scope guards inherited ----------

def test_export_scope_guards(ctx):
    async def go():
        # cashier/kitchen 403 on any export
        for tok in (ctx.cashier_a, ctx.kitchen_a):
            assert (await ctx.client.get("/api/reports/pnl/export.xlsx", headers=auth(tok))).status_code == 403
            assert (await ctx.client.get("/api/reports/items/export.pdf", headers=auth(tok))).status_code == 403
        # area cross-cluster 403
        r = await ctx.client.get(f"/api/reports/pnl/export.xlsx?store_ids={ctx.store_z}", headers=auth(ctx.area_ab))
        assert r.status_code == 403
        # store_manager other store 403
        r2 = await ctx.client.get(f"/api/reports/pnl/export.pdf?store_ids={ctx.store_b}", headers=auth(ctx.mgr_a))
        assert r2.status_code == 403
    run(go())


def test_consolidated_export_hq_only(ctx):
    async def go():
        ok = await ctx.client.get("/api/reports/consolidated/export.xlsx", headers=auth(ctx.hq))
        assert ok.status_code == 200 and ok.content[:2] == b"PK"
        for tok in (ctx.area_ab, ctx.mgr_a, ctx.cashier_a):
            r = await ctx.client.get("/api/reports/consolidated/export.xlsx", headers=auth(tok))
            assert r.status_code == 403
    run(go())


def test_inventory_export_cost_restricted(ctx):
    async def go():
        ok = await ctx.client.get(f"/api/reports/inventory/export.xlsx?store_ids={ctx.store_a}", headers=auth(ctx.mgr_a))
        assert ok.status_code == 200
        for tok in (ctx.cashier_a, ctx.kitchen_a):
            r = await ctx.client.get(f"/api/reports/inventory/export.xlsx?store_ids={ctx.store_a}", headers=auth(tok))
            assert r.status_code == 403
    run(go())
